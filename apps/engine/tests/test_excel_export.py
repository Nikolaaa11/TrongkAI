"""Tests del export Excel del Plan 5 Años."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from trongkai_engine.excel_export import export_plan_to_excel
from trongkai_engine.plan_builder import build_plan


@pytest.fixture
def temp_xlsx(tmp_path: Path) -> Path:
    return tmp_path / "plan.xlsx"


def test_export_genera_archivo(temp_xlsx):
    plan = build_plan()
    path = export_plan_to_excel(plan, temp_xlsx)
    assert path.exists()
    assert path.stat().st_size > 5000


def test_export_tiene_4_hojas(temp_xlsx):
    plan = build_plan()
    export_plan_to_excel(plan, temp_xlsx)
    wb = openpyxl.load_workbook(temp_xlsx)
    assert "Supuestos" in wb.sheetnames
    assert "EERR_Mensual" in wb.sheetnames
    assert "KPIs" in wb.sheetnames
    assert "Resumen_Anual" in wb.sheetnames


def test_eerr_tiene_60_filas_de_datos(temp_xlsx):
    plan = build_plan()
    export_plan_to_excel(plan, temp_xlsx)
    wb = openpyxl.load_workbook(temp_xlsx)
    ws = wb["EERR_Mensual"]
    # Encabezado fila 3 + 60 meses + total
    primer_mes = ws.cell(row=4, column=1).value
    mes_60 = ws.cell(row=63, column=1).value
    assert primer_mes == 1
    assert mes_60 == 60


def test_kpis_se_escribe(temp_xlsx):
    plan = build_plan()
    export_plan_to_excel(plan, temp_xlsx)
    wb = openpyxl.load_workbook(temp_xlsx)
    ws = wb["KPIs"]
    rows_text = [ws.cell(row=r, column=1).value for r in range(1, 10)]
    assert any("TIR" in (t or "") for t in rows_text)
    assert any("VAN" in (t or "") for t in rows_text)
    assert any("Payback" in (t or "") for t in rows_text)


def test_formato_money_aplica_parentesis_a_negativos(temp_xlsx):
    """Regla §4 M3: negativos en paréntesis. El formato debe contener '('
    para que Excel muestre así los valores negativos en columnas de plata."""
    plan = build_plan()
    export_plan_to_excel(plan, temp_xlsx)
    wb = openpyxl.load_workbook(temp_xlsx)
    ws = wb["EERR_Mensual"]
    # Columnas 2..11 son monetarias en filas de datos (4..63)
    for col in range(2, 12):
        fmt = ws.cell(row=4, column=col).number_format
        assert "(" in fmt and ")" in fmt, f"col {col} sin paréntesis para negativos: {fmt}"


def test_resumen_anual_tiene_5_anios(temp_xlsx):
    plan = build_plan()
    export_plan_to_excel(plan, temp_xlsx)
    wb = openpyxl.load_workbook(temp_xlsx)
    ws = wb["Resumen_Anual"]
    labels = [ws.cell(row=r, column=1).value for r in range(4, 9)]
    assert labels == [f"Año {i+1}" for i in range(5)]
    # Margin format en columna 4 debe ser porcentaje
    assert ws.cell(row=4, column=4).number_format == "0.00%"


def test_eerr_fila_total_tiene_formulas_sum(temp_xlsx):
    """La fila TOTAL debe contener fórmulas =SUM(...), no valores hardcodeados.
    Verifica que la integridad del libro no tenga #REF! ni hardcodes."""
    plan = build_plan()
    export_plan_to_excel(plan, temp_xlsx)
    # Cargamos con keep_links=False y data_only=False para ver fórmulas
    wb = openpyxl.load_workbook(temp_xlsx, data_only=False)
    ws = wb["EERR_Mensual"]
    # Buscar fila TOTAL
    total_row = None
    for r in range(60, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "TOTAL":
            total_row = r
            break
    assert total_row is not None
    # Columnas 2..11 deben tener fórmulas SUM
    for col in range(2, 12):
        val = ws.cell(row=total_row, column=col).value
        assert isinstance(val, str) and val.startswith("=SUM("), (
            f"col {col} en fila TOTAL no es fórmula: {val}"
        )
        assert "#REF" not in val


def test_supuestos_inputs_marcados_azul(temp_xlsx):
    """Los valores de inputs en Supuestos deben tener fill azul (DCE6F1).
    Cobertura semántica del estilo AZUL_INPUT."""
    plan = build_plan()
    export_plan_to_excel(plan, temp_xlsx)
    wb = openpyxl.load_workbook(temp_xlsx)
    ws = wb["Supuestos"]
    # Fila 4 = primer input (Volumen total anual)
    cell = ws.cell(row=4, column=2)
    fill_color = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
    # openpyxl puede devolver "00DCE6F1" (con alpha) o "DCE6F1"
    assert fill_color and "DCE6F1" in fill_color, f"fill no es azul: {fill_color}"
