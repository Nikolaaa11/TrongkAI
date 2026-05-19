"""Export del Plan 5 Años a Excel formato directorio.

Reglas (SUPER_PROMPT §4 M3):
- Hoja Supuestos con celdas azules para inputs.
- EERR mensual 60 meses, fórmulas en negro, links en verde.
- KPIs y Tornado.
- Negativos en paréntesis.
- Sin #REF!
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .plan_builder import ResumenPlan

if TYPE_CHECKING:
    from pathlib import Path

AZUL_INPUT = PatternFill("solid", fgColor="DCE6F1")
VERDE_LINK = PatternFill("solid", fgColor="EAF2D3")
NEGRO_FORMULA = PatternFill("solid", fgColor="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="3F4A2B")
HEADER_FONT = Font(bold=True, color="F5F1E8")
TOTAL_FONT = Font(bold=True)
MONEY_FMT = '"$"#,##0;("$"#,##0);"-"'


def export_plan_to_excel(plan: ResumenPlan, path: Path) -> Path:
    wb = Workbook()
    _write_supuestos(wb, plan)
    _write_eerr(wb, plan)
    _write_kpis(wb, plan)
    _write_resumen_anual(wb, plan)

    # Eliminar hoja default
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _write_supuestos(wb: Workbook, plan: ResumenPlan) -> None:
    ws = wb.create_sheet("Supuestos", 0)
    ws.append(["Trongkai — Supuestos del Plan 5 Años", "", "", ""])
    ws["A1"].font = Font(size=14, bold=True, color="3F4A2B")
    ws.append([])
    ws.append(["Variable", "Valor", "Unidad", "Notas"])
    for c in ws[3]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    rows = [
        ("Volumen total anual", plan.parametros.volumen_total_ton_ano, "ton/año", "Cuota contractual"),
        ("WACC anual", plan.parametros.wacc_anual, "fracción", "PD — sin firma directorio"),
        ("OpEx mensual", plan.parametros.opex_mensual_clp, "CLP/mes", "MO + mantención + admin"),
        ("Maquilas mensual", plan.parametros.maquilas_mensual_clp, "CLP/mes", "Ingreso accesorio"),
        ("Transferencia tec. anual (años 1-2)", plan.parametros.transferencia_tec_anual_clp, "CLP/año", "Patines CORFO"),
        ("Costo MMPP promedio", plan.parametros.costo_mmpp_clp_kg, "CLP/kg input", "Flete + pago al proveedor neto"),
    ]
    for var, valor, unidad, nota in rows:
        ws.append([var, valor, unidad, nota])
        ws.cell(row=ws.max_row, column=2).fill = AZUL_INPUT
        ws.cell(row=ws.max_row, column=2).number_format = MONEY_FMT if "CLP" in unidad else "0.000"

    ws.append([])
    ws.append(["Precios de venta por SKU"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["SKU", "CLP/kg", "Marca", ""])
    for c in ws[ws.max_row]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for sku, precio in plan.parametros.precios_clp_kg.items():
        ws.append([sku, precio, "", ""])
        ws.cell(row=ws.max_row, column=2).fill = AZUL_INPUT
        ws.cell(row=ws.max_row, column=2).number_format = MONEY_FMT

    ws.append([])
    ws.append(["Rendimiento por MMPP"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["MMPP", "Rendimiento", "", ""])
    for c in ws[ws.max_row]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for mmpp, rend in plan.parametros.rendimiento_por_mmpp.items():
        ws.append([mmpp, rend, "", ""])
        ws.cell(row=ws.max_row, column=2).fill = AZUL_INPUT
        ws.cell(row=ws.max_row, column=2).number_format = "0.00%"

    _autosize(ws)


def _write_eerr(wb: Workbook, plan: ResumenPlan) -> None:
    ws = wb.create_sheet("EERR_Mensual", 1)
    ws.append(["EERR mensual — 60 meses"])
    ws["A1"].font = Font(size=14, bold=True, color="3F4A2B")
    ws.append([])
    headers = [
        "Mes",
        "Ingresos ventas",
        "Maquilas",
        "Recepción residuos",
        "Transferencia tec.",
        "Ingresos totales",
        "Costos directos",
        "Gastos fijos",
        "EBITDA",
        "CapEx",
        "Flujo neto",
    ]
    ws.append(headers)
    for c in ws[3]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")

    for f in plan.flujos:
        ws.append([
            f.mes,
            f.ingresos_ventas,
            f.ingresos_maquilas,
            f.ingresos_recepcion,
            f.ingresos_transferencia_tec,
            f.ingresos_totales,
            f.costos_directos,
            f.gastos_fijos,
            f.ebitda,
            f.capex_periodo,
            f.flujo_neto,
        ])

    # Aplicar formato monetario a todas las columnas de plata
    last_row = ws.max_row
    for col in range(2, 12):
        for r in range(4, last_row + 1):
            ws.cell(row=r, column=col).number_format = MONEY_FMT
        # Las columnas calculadas (totales y EBITDA y flujo neto) llevan verde (link)
        if col in (6, 9, 11):  # ingresos_totales, EBITDA, flujo_neto
            for r in range(4, last_row + 1):
                ws.cell(row=r, column=col).fill = VERDE_LINK

    # Fila de totales
    ws.append([])
    ws.append([
        "TOTAL",
        f"=SUM(B4:B{last_row})",
        f"=SUM(C4:C{last_row})",
        f"=SUM(D4:D{last_row})",
        f"=SUM(E4:E{last_row})",
        f"=SUM(F4:F{last_row})",
        f"=SUM(G4:G{last_row})",
        f"=SUM(H4:H{last_row})",
        f"=SUM(I4:I{last_row})",
        f"=SUM(J4:J{last_row})",
        f"=SUM(K4:K{last_row})",
    ])
    total_row = ws.max_row
    for c in ws[total_row]:
        c.font = TOTAL_FONT
        if c.column > 1:
            c.number_format = MONEY_FMT
    _autosize(ws)


def _write_kpis(wb: Workbook, plan: ResumenPlan) -> None:
    ws = wb.create_sheet("KPIs", 2)
    ws.append(["KPIs financieros"])
    ws["A1"].font = Font(size=14, bold=True, color="3F4A2B")
    ws.append([])
    rows = [
        ("TIR proyecto anual", plan.kpis.tir_proyecto_anual, "0.00%"),
        ("VAN", plan.kpis.van, MONEY_FMT),
        ("Payback descontado (meses)", plan.kpis.payback_meses or "—", "0"),
        ("EBITDA margin promedio", plan.kpis.ebitda_margin_promedio, "0.00%"),
        ("Ratio CapEx/Ventas", plan.kpis.ratio_capex_ventas, "0.00%"),
    ]
    for label, val, fmt in rows:
        ws.append([label, val])
        ws.cell(row=ws.max_row, column=2).number_format = fmt
        ws.cell(row=ws.max_row, column=2).fill = VERDE_LINK
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    _autosize(ws)


def _write_resumen_anual(wb: Workbook, plan: ResumenPlan) -> None:
    ws = wb.create_sheet("Resumen_Anual", 3)
    ws.append(["Resumen anual"])
    ws["A1"].font = Font(size=14, bold=True, color="3F4A2B")
    ws.append([])
    ws.append(["Año", "Ingresos", "EBITDA", "EBITDA Margin", "CapEx"])
    for c in ws[3]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    for i in range(5):
        ingreso = plan.ingresos_anuales[i]
        ebitda = plan.ebitda_anuales[i]
        margin = ebitda / ingreso if ingreso else 0
        capex = plan.capex_anuales[i]
        ws.append([f"Año {i+1}", ingreso, ebitda, margin, capex])
        for col in (2, 3, 5):
            ws.cell(row=ws.max_row, column=col).number_format = MONEY_FMT
            ws.cell(row=ws.max_row, column=col).fill = VERDE_LINK
        ws.cell(row=ws.max_row, column=4).number_format = "0.00%"
    _autosize(ws)


def _autosize(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 35)
