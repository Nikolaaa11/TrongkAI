"""Seed inicial de la DB Trongkai desde los 3 Excels del cliente.

Modos:
  --dry-run  : sólo vuelca payload a scripts/seed_dryrun.json (default si no hay DATABASE_URL)
  --apply    : conecta a Postgres y hace UPSERT (requiere DATABASE_URL y psycopg)

Lee:
- contexto/Info_Plan_5_anos_Estructura_A.xlsx (proveedores + base rendimiento)
- contexto/Cuadro_PPTO_Variables_PD_Plan_5_Anos_A.xlsx (matriz PD/OK)

Idempotente: re-ejecutar no duplica registros.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import asdict, dataclass, field
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
CTX = ROOT / "contexto"

INFO_FILE = CTX / "Info_Plan_5_anos_Estructura_A.xlsx"
CUADRO_FILE = CTX / "Cuadro_PPTO_Variables_PD_Plan_5_Anos_A.xlsx"


@dataclass
class SeedPayload:
    materias_primas: list[dict] = field(default_factory=list)
    suppliers: list[dict] = field(default_factory=list)
    productos: list[dict] = field(default_factory=list)
    supuestos: list[dict] = field(default_factory=list)
    capacidades: list[dict] = field(default_factory=list)


MMPP_CATALOGO = [
    dict(
        codigo="ALPERUJO", nombre="Alperujo de olivas",
        temporadaInicioMes=4, temporadaFinMes=6,
        humedadInicialPct=0.65, materiaSolidaPct=0.35,
        aceiteExtraiblePct=0.02,
        notas="Subproducto almazaras de oliva. Objetivo aceite 2% con Opticept (competencia 1.3-1.45).",
    ),
    dict(
        codigo="TOMASA", nombre="Tomasa (subproducto tomate)",
        temporadaInicioMes=1, temporadaFinMes=3,
        humedadInicialPct=0.82, materiaSolidaPct=0.18,
        licopenoPct=0.001, pectinaPct=0.002,
        notas="Humedad ~82% — peor MS. Tiempo descomposición PD (estimado 3h).",
    ),
    dict(
        codigo="POMASA", nombre="Pomasa (manzana/pera)",
        temporadaInicioMes=3, temporadaFinMes=5,
        humedadInicialPct=0.80, materiaSolidaPct=0.20,
        pectinaPct=0.003,
    ),
    dict(
        codigo="ORUJO_UVA", nombre="Orujo de uva",
        temporadaInicioMes=4, temporadaFinMes=6,
        humedadInicialPct=None, materiaSolidaPct=None,
        notas="Humedad variable — pendiente medición.",
    ),
    dict(
        codigo="LEVADURA", nombre="Levadura (cerveza/vino)",
        temporadaInicioMes=1, temporadaFinMes=12,
        notas="Año corrido. PTEC: proteína unicelular.",
    ),
]


def parse_suppliers(ws) -> list[dict]:
    """Extrae los 4 oliveros con datos del Excel + placeholders."""
    suppliers = []
    for row in range(5, 15):
        nombre = ws.cell(row=row, column=2).value
        if not nombre:
            continue
        dist = ws.cell(row=row, column=3).value
        flete_1800 = ws.cell(row=row, column=4).value or 0
        flete_2100 = ws.cell(row=row, column=5).value or 0
        flete_2500 = ws.cell(row=row, column=6).value or 0
        if flete_1800 and not isinstance(flete_1800, str):
            tarifa = 1800
        elif flete_2100 and not isinstance(flete_2100, str):
            tarifa = 2100
        elif flete_2500 and not isinstance(flete_2500, str):
            tarifa = 2500
        else:
            tarifa = 0
        pago_recepcion = ws.cell(row=row, column=8).value or 0
        volumen = ws.cell(row=row, column=9).value or 0
        cap_camion = ws.cell(row=row, column=10).value or 22.5
        caso_str = ws.cell(row=row, column=11).value or ""
        caso_logistico = {
            "Caso 1": "CASO_1",
            "Caso 2": "CASO_2",
            "Caso 3 ": "CASO_3",
            "Caso 3": "CASO_3",
            "Caso 4": "CASO_4",
        }.get(caso_str.strip() if isinstance(caso_str, str) else "", "CASO_3")

        if dist in (None, "") or not isinstance(dist, (int, float)):
            status = "PROSPECT"
            dist_val = 0.0
        else:
            status = "ACTIVO"
            dist_val = float(dist)

        suppliers.append(
            dict(
                nombre=str(nombre).strip(),
                mmppCodigo="ALPERUJO",
                distanciaKm=dist_val,
                tarifaFleteClpKm=float(tarifa),
                casoLogistico=caso_logistico,
                pagoRecepcionClpKg=float(pago_recepcion) if isinstance(pago_recepcion, (int, float)) else 0.0,
                volumenAnualComprometidoTon=float(volumen) if isinstance(volumen, (int, float)) else 0.0,
                capacidadCamionTon=float(cap_camion) if isinstance(cap_camion, (int, float)) else 22.5,
                status=status,
            )
        )
    return suppliers


# Marcas según ADR-009. Algunos SKUs son cross-brand (ej: HARINA_TOMASA aparece en Feed y Food)
# pero en DB cada Producto vive con UNA marca primaria. Los cross-brand quedan duplicados
# como SKUs separados con sufijo si el negocio lo requiere — por ahora marca primaria.
PRODUCTOS_CATALOGO = [
    # (codigo, nombre, mmpp_origen, tipo, marca, ano_lanzamiento)
    ("HARINA_ALPERUJO", "Harina de Alperujo",       "ALPERUJO",  "BASE",     "FOOD", 1),
    ("ACEITE_ALPERUJO", "Aceite de Alperujo",       "ALPERUJO",  "BASE",     "FOOD", 1),
    ("HARINA_ORUJO",    "Harina de Orujo",          "ORUJO_UVA", "BASE",     "FEED", 1),
    ("HARINA_TOMASA",   "Harina de Tomasa",         "TOMASA",    "BASE",     "FOOD", 1),
    ("HARINA_POMASA",   "Harina de Pomasa",         "POMASA",    "BASE",     "FEED", 1),
    ("PECTINA",         "Pectina",                  "POMASA",    "AGREGADO", "FOOD", 2),
    ("LICOPENO",        "Licopeno",                 "TOMASA",    "AGREGADO", "FOOD", 2),
    ("PROTEINA_UNICEL", "Proteína Unicelular",      "LEVADURA",  "PTEC",     "FEED", 3),
    ("ANTIOXIDANTE",    "Antioxidante",             "ALPERUJO",  "PTEC",     "FEED", 3),
    ("AGLOMERANTE",     "Aglomerante",              None,        "PTEC",     "FEED", 4),
    ("UMAMI",           "Umami",                    None,        "PTEC",     "FOOD", 5),
    ("ACEITE_ORUJO_UVA","Aceite de Orujo (eventual)","ORUJO_UVA","BASE",     "FOOD", 2),
]


def parse_supuestos_cuadro(ws) -> list[dict]:
    """Extrae los PD/OK/OK* de la matriz Datos x Plan 5 años."""
    out = []
    sku_headers = {}
    for col in range(3, 16):
        h = ws.cell(row=4, column=col).value
        if h:
            sku_headers[col] = str(h).strip()

    estado_map = {
        "PD": "PD",
        "OK": "OK_PROVISORIO",
        "OK*": "OK_PROVISORIO",
        "PD/2": "PD",
        "OK*/PD": "PD",
    }

    for row in range(4, 24):
        variable = ws.cell(row=row, column=2).value
        if not variable:
            continue
        for col, sku in sku_headers.items():
            v = ws.cell(row=row, column=col).value
            if v is None or v == "":
                continue
            v_str = str(v).strip()
            estado = estado_map.get(v_str, "NO_APLICA")
            clave = f"{str(variable).lower().replace(' ', '_').strip()}.{sku.lower().replace('.', '').replace(' ', '_')}"
            out.append(
                dict(
                    clave=clave[:200],
                    valorActual=None,
                    unidad=None,
                    fuente="Cuadro_PPTO_Variables_PD_Plan_5_Anos_A.xlsx",
                    estado=estado,
                    sensibilidad="MEDIA",
                    owner="Matías" if any(s in str(variable) for s in ("Volumen", "Costo Transporte")) else "Jaime",
                    tag=None,
                )
            )
    return out


def parse_capacidades_pd() -> list[dict]:
    etapas = [
        "RECEPCION", "ALIMENTACION", "HOMOG_1", "PEF",
        "PRENSADO_MECANICO", "TRICANTER", "EXTRACCION",
        "SECADO", "HOMOG_2", "ENSACADO",
    ]
    return [
        dict(
            etapa=e,
            capacidadTonHora=None,
            tiempoResidenciaH=None,
            opcional=(e in ("PEF", "TRICANTER", "EXTRACCION")),
        )
        for e in etapas
    ]


def build_payload() -> SeedPayload:
    payload = SeedPayload()
    payload.materias_primas = MMPP_CATALOGO

    wb_info = openpyxl.load_workbook(INFO_FILE, data_only=True)
    ws_prov = wb_info["IngresosCostos Proveedores"]
    payload.suppliers = parse_suppliers(ws_prov)

    payload.productos = [
        dict(codigo=c, nombre=n, mmppOrigen=m, tipo=t, marca=mk, anoLanzamiento=a)
        for c, n, m, t, mk, a in PRODUCTOS_CATALOGO
    ]

    wb_cuadro = openpyxl.load_workbook(CUADRO_FILE, data_only=True)
    ws_datos = wb_cuadro["Datos x Plan 5 años"]
    payload.supuestos = parse_supuestos_cuadro(ws_datos)

    payload.capacidades = parse_capacidades_pd()
    return payload


def apply_to_db(payload: SeedPayload) -> dict:
    """Inserta payload en Postgres. Devuelve contadores."""
    sys.path.insert(0, str(ROOT / "apps" / "engine"))
    from trongkai_engine import repository as repo  # noqa: WPS433

    counts = {}
    with repo.connect() as conn:
        counts["materias_primas"] = repo.upsert_materias_primas(conn, payload.materias_primas)
        counts["suppliers"] = repo.upsert_suppliers(conn, payload.suppliers)
        counts["productos"] = repo.upsert_productos(conn, payload.productos)
        counts["supuestos"] = repo.upsert_supuestos(conn, payload.supuestos)
        counts["capacidades"] = repo.upsert_capacidades(conn, payload.capacidades)
        conn.commit()
        stats = repo.stats_resumen(conn)
    counts["db_stats"] = dict(stats)
    return counts


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="No conectar a DB; volcar JSON.")
    parser.add_argument("--apply", action="store_true", help="Aplicar contra Postgres.")
    args = parser.parse_args()

    payload = build_payload()

    use_db = args.apply or (not args.dry_run and "DATABASE_URL" in os.environ)

    if not use_db:
        out_path = ROOT / "scripts" / "seed_dryrun.json"
        out_path.write_text(json.dumps(asdict(payload), indent=2, ensure_ascii=False, default=str), encoding="utf-8")
        print(f"[dry-run] Payload escrito en {out_path}")
        print(f"  - MMPP:        {len(payload.materias_primas)}")
        activos = sum(1 for s in payload.suppliers if s['status'] == 'ACTIVO')
        print(f"  - Suppliers:   {len(payload.suppliers)} ({activos} activos)")
        print(f"  - Productos:   {len(payload.productos)}")
        feed = sum(1 for p in payload.productos if p['marca'] == 'FEED')
        food = sum(1 for p in payload.productos if p['marca'] == 'FOOD')
        print(f"     Feed:       {feed}")
        print(f"     Food:       {food}")
        print(f"  - Supuestos:   {len(payload.supuestos)}")
        print(f"  - Capacidades: {len(payload.capacidades)}")
        return 0

    print("[apply] Conectando a Postgres...")
    try:
        result = apply_to_db(payload)
    except Exception as exc:
        print(f"[apply] FAILED: {exc}", file=sys.stderr)
        return 2

    print("[apply] OK")
    for k, v in result.items():
        print(f"  {k}: {v}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
