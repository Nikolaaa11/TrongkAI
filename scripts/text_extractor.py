"""Text Extractor - extrae texto de archivos del inbox para clasificación más inteligente.

Soporta:
- .pdf       (pypdf - fallback graceful si no está instalado)
- .xlsx/.xls (openpyxl)
- .docx      (python-docx)
- .txt/.md/.csv/.tsv (texto plano)
- Otros: skip

Plus: extrae entidades (montos, fechas, contrapartes) con regex.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

MAX_CHARS = 5000  # No saturar el clasificador


def extraer_texto(path: Path) -> dict[str, Any]:
    """Extrae texto + entidades de un archivo. Devuelve dict con 'texto', 'entidades', 'error'."""
    ext = path.suffix.lower()
    result: dict[str, Any] = {
        "texto": "",
        "entidades": {},
        "metodo": "skipped",
        "error": None,
    }

    try:
        if ext == ".pdf":
            result["texto"] = _extract_pdf(path)
            result["metodo"] = "pypdf"
        elif ext in (".xlsx", ".xls", ".xlsm"):
            result["texto"] = _extract_excel(path)
            result["metodo"] = "openpyxl"
        elif ext == ".docx":
            result["texto"] = _extract_docx(path)
            result["metodo"] = "python-docx"
        elif ext in (".txt", ".md", ".csv", ".tsv"):
            result["texto"] = path.read_text(encoding="utf-8", errors="replace")
            result["metodo"] = "text-plain"
        else:
            result["metodo"] = "unsupported-extension"
            return result
    except Exception as e:
        result["error"] = str(e)
        return result

    # Truncar y extraer entidades
    if len(result["texto"]) > MAX_CHARS:
        result["texto"] = result["texto"][:MAX_CHARS] + f"\n... [truncado, {len(result['texto'])} chars total]"

    result["entidades"] = _extraer_entidades(result["texto"])
    return result


def _extract_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        return "(pypdf no instalado - pip install pypdf)"
    reader = PdfReader(str(path))
    chunks = []
    for page in reader.pages[:10]:  # Max 10 pages
        chunks.append(page.extract_text() or "")
    return "\n".join(chunks)


def _extract_excel(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "(openpyxl no instalado)"
    wb = load_workbook(str(path), data_only=True, read_only=True)
    chunks = []
    for sheet_name in wb.sheetnames[:5]:  # Max 5 sheets
        ws = wb[sheet_name]
        chunks.append(f"=== {sheet_name} ===")
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx > 50:  # Max 50 rows per sheet
                chunks.append("... [more rows]")
                break
            vals = [str(c) for c in row if c is not None]
            if vals:
                chunks.append(" | ".join(vals))
    return "\n".join(chunks)


def _extract_docx(path: Path) -> str:
    try:
        from docx import Document
    except ImportError:
        return "(python-docx no instalado - pip install python-docx)"
    doc = Document(str(path))
    chunks = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(chunks)


# ============================================================================
# Entity extraction (regex-based, sin deps)
# ============================================================================

REGEX_MONTO_USD = re.compile(r"(?:USD|US\$|\$US)\s?([\d.,]+)\s?(M|millones|MM|K|mil)?", re.IGNORECASE)
REGEX_MONTO_CLP = re.compile(r"(?:CLP|\$)\s?([\d.,]+)\s?(M|MM|millones|B|billones|K|mil)?", re.IGNORECASE)
REGEX_FECHA = re.compile(r"\b(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\b|\b(\d{4})[/-](\d{1,2})[/-](\d{1,2})\b")
REGEX_TONS = re.compile(r"(\d+(?:\.\d+)?)\s?(?:ton(?:eladas)?|tn|kg)", re.IGNORECASE)
REGEX_PCT = re.compile(r"(\d+(?:\.\d+)?)\s?%")
REGEX_EMPRESA = re.compile(r"\b(?:S\.A\.|SpA|Ltda\.|Limitada|SAS|S\.R\.L\.)\b")
REGEX_RUT = re.compile(r"\b\d{1,3}(?:\.\d{3}){0,2}-[\dKk]\b")

# Keywords clave para Trongkai
KEYWORDS = {
    "mmpp": ["alperujo", "tomasa", "pomasa", "orujo", "olivar", "tomatera", "vinícola", "vinicola"],
    "skus": ["harina", "aceite", "pectina", "licopeno", "proteína", "proteina", "antioxidante", "aglomerante", "umami"],
    "comercial": ["loi", "carta de intención", "carta intencion", "letter of intent", "cotización", "cotizacion", "purchase order"],
    "financiero": ["tir", "van", "ebitda", "capex", "opex", "dscr", "term sheet", "wacc"],
    "esg": ["lca", "huella carbono", "b-corp", "haccp", "gmp", "iso 14001", "rep", "ods"],
}


def _extraer_entidades(texto: str) -> dict[str, list]:
    """Extrae entidades estructuradas desde el texto."""
    if not texto:
        return {}

    montos_usd = REGEX_MONTO_USD.findall(texto)[:5]
    montos_clp = REGEX_MONTO_CLP.findall(texto)[:5]
    fechas = REGEX_FECHA.findall(texto)[:10]
    tons = REGEX_TONS.findall(texto)[:5]
    pcts = REGEX_PCT.findall(texto)[:10]
    empresas = REGEX_EMPRESA.findall(texto)[:5]
    ruts = REGEX_RUT.findall(texto)[:3]

    # Keywords detectados
    texto_lower = texto.lower()
    keywords_hit = {}
    for cat, kws in KEYWORDS.items():
        hits = [k for k in kws if k in texto_lower]
        if hits:
            keywords_hit[cat] = hits

    return {
        "montos_usd": [f"{m[0]}{m[1] or ''}" for m in montos_usd],
        "montos_clp": [f"{m[0]}{m[1] or ''}" for m in montos_clp],
        "fechas_detectadas": fechas[:5],
        "toneladas": tons,
        "porcentajes": pcts[:5],
        "tiene_estructura_legal_chile": len(empresas) > 0 or len(ruts) > 0,
        "ruts_chilenos": ruts,
        "empresas_tipo": list(set(empresas)),
        "keywords": keywords_hit,
    }


def resumen_entidades(entidades: dict) -> str:
    """Resumen compacto para audit log."""
    if not entidades:
        return ""
    parts = []
    if entidades.get("montos_usd"):
        parts.append(f"USD: {', '.join(entidades['montos_usd'][:3])}")
    if entidades.get("montos_clp"):
        parts.append(f"CLP: {', '.join(entidades['montos_clp'][:3])}")
    if entidades.get("toneladas"):
        parts.append(f"ton: {', '.join(entidades['toneladas'][:3])}")
    if entidades.get("fechas_detectadas"):
        parts.append(f"{len(entidades['fechas_detectadas'])} fechas")
    if entidades.get("keywords"):
        for cat, hits in entidades["keywords"].items():
            parts.append(f"{cat}:{','.join(hits[:2])}")
    return " | ".join(parts) if parts else ""


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python text_extractor.py <archivo>")
        sys.exit(1)
    p = Path(sys.argv[1])
    if not p.exists():
        print(f"No existe: {p}")
        sys.exit(1)
    result = extraer_texto(p)
    print(f"Método: {result['metodo']}")
    if result["error"]:
        print(f"Error: {result['error']}")
    print(f"Texto extraído ({len(result['texto'])} chars):")
    print(result["texto"][:1000])
    print()
    print("Entidades:")
    import json
    print(json.dumps(result["entidades"], indent=2, ensure_ascii=False))
    print()
    print(f"Resumen: {resumen_entidades(result['entidades'])}")
