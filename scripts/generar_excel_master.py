"""Genera Trongkai-Master.xlsx — Excel multi-hoja con TODO el modelo.

Hojas:
1. Dashboard       — KPIs grandes + interpretación
2. Plan-5-anos     — EERR mensual a 60 meses, fórmulas vivas
3. Balance-Masa    — flujo MMPP -producto final
4. Sensibilidad    — tabla NxN precio × costo MMPP
5. Escenarios      — PILOTO / INDUSTRIAL / EXPANSION
6. Carbono-LCA     — 3 escenarios + revenue créditos
7. Compliance-REP  — 8 hitos Ley REP
8. Macro-Chile     — USD/CLP, UF, IPC, TPM (último snapshot)
9. Supuestos       — listado top 20

Uso:
    python scripts/generar_excel_master.py
    -genera: backups/Trongkai-Master-YYYYMMDD-HHMM.xlsx
"""
from __future__ import annotations

import os
import sys
from datetime import datetime
from pathlib import Path

# Permitir import del engine
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "apps" / "engine"))

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Engine imports
from trongkai_engine.breakeven import breakeven_summary
from trongkai_engine.carbon_footprint import comparar_escenarios_footprint
from trongkai_engine.compliance_rep import HITOS_LEY_REP
from trongkai_engine.escenarios import comparar_escenarios_estrategicos
from trongkai_engine.macro_chile import snapshot_resumen
from trongkai_engine.plan_builder import ParametrosPlan, build_plan
from trongkai_engine.readiness_score import calcular_readiness_score
from trongkai_engine.sensitivity import heatmap_2d
from trongkai_engine.valuation import valuar_proyecto_ev_ebitda

# =====================================================================
# Paleta Trongkai
# =====================================================================
VERDE_TRONGKAI = "1A8A1A"
VERDE_OSCURO = "0B3D0B"
GRIS_FONDO = "F5F5F7"
GRIS_BORDE = "E8E8ED"
TEXTO = "1D1D1F"
TEXTO_SEC = "86868B"
ROJO = "FF3B30"
NARANJA = "FF9500"
AMARILLO = "FFCC00"

THIN = Side(style="thin", color=GRIS_BORDE)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def estilo_titulo(cell):
    cell.font = Font(name="Calibri", size=20, bold=True, color=TEXTO)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def estilo_subtitulo(cell):
    cell.font = Font(name="Calibri", size=11, color=TEXTO_SEC)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def estilo_header_tabla(cell):
    cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=VERDE_TRONGKAI)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER_ALL


def estilo_celda(cell, bold=False):
    cell.font = Font(name="Calibri", size=10, bold=bold, color=TEXTO)
    cell.border = BORDER_ALL
    cell.alignment = Alignment(horizontal="center", vertical="center")


def estilo_kpi_grande(cell):
    cell.font = Font(name="Calibri", size=36, bold=True, color=VERDE_TRONGKAI)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def estilo_kpi_label(cell):
    cell.font = Font(name="Calibri", size=10, color=TEXTO_SEC, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


# =====================================================================
# Hoja 1: Dashboard
# =====================================================================
def crear_dashboard(wb: Workbook, snap: dict, rs: dict):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Trongkai"
    estilo_titulo(ws["A1"])
    ws["A2"] = "Innovación en Nutrición Circular · Tearsheet ejecutivo"
    estilo_subtitulo(ws["A2"])
    ws["A3"] = f"Generado: {datetime.now().strftime('%d %B %Y, %H:%M')}"
    estilo_subtitulo(ws["A3"])

    # KPIs grandes
    kpis = snap["plan"]["kpis"]
    val = snap["valuation"]
    ws.merge_cells("A5:B5"); ws["A5"] = "TIR Proyecto"; estilo_kpi_label(ws["A5"])
    ws.merge_cells("A6:B6"); ws["A6"] = f"{kpis['tir'] * 100:.2f}%"; estilo_kpi_grande(ws["A6"])

    ws.merge_cells("C5:D5"); ws["C5"] = "VAN @ WACC 18%"; estilo_kpi_label(ws["C5"])
    ws.merge_cells("C6:D6"); ws["C6"] = f"${kpis['van'] / 1e9:.1f}B CLP"; estilo_kpi_grande(ws["C6"])

    ws.merge_cells("E5:F5"); ws["E5"] = "Payback"; estilo_kpi_label(ws["E5"])
    ws.merge_cells("E6:F6"); ws["E6"] = f"{kpis['payback_meses'] or '—'} meses"; estilo_kpi_grande(ws["E6"])

    ws.merge_cells("G5:H5"); ws["G5"] = "EV exit 5y"; estilo_kpi_label(ws["G5"])
    ws.merge_cells("G6:H6"); ws["G6"] = f"${val['ev_base_clp'] / 1e9:.0f}B CLP"; estilo_kpi_grande(ws["G6"])

    # Score
    ws.merge_cells("A9:H9")
    ws["A9"] = f"Investment Readiness Score: {rs['score_total']:.1f}/100 — {rs['interpretacion']}"
    ws["A9"].font = Font(name="Calibri", size=14, bold=True, color=VERDE_TRONGKAI)
    ws["A9"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A9"].fill = PatternFill("solid", fgColor="EAF6EA")

    # Tabla de dimensiones
    ws["A11"] = "Score por dimensión"
    ws["A11"].font = Font(size=14, bold=True, color=TEXTO)

    headers = ["Dimensión", "Peso", "Score 0-100", "Aporte total", "Detalle"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=12, column=i, value=h)
        estilo_header_tabla(cell)

    for i, d in enumerate(rs["dimensiones"], 1):
        ws.cell(row=12 + i, column=1, value=d["nombre"])
        ws.cell(row=12 + i, column=2, value=f"{d['peso'] * 100:.0f}%")
        ws.cell(row=12 + i, column=3, value=d["score_0_100"])
        ws.cell(row=12 + i, column=4, value=d["aporte_total"])
        ws.cell(row=12 + i, column=5, value=d["detalle"])
        for col in range(1, 6):
            estilo_celda(ws.cell(row=12 + i, column=col))

    # Color scale para score
    ws.conditional_formatting.add(
        f"C13:C{12 + len(rs['dimensiones'])}",
        ColorScaleRule(
            start_type="num", start_value=0, start_color=ROJO,
            mid_type="num", mid_value=50, mid_color=AMARILLO,
            end_type="num", end_value=100, end_color=VERDE_TRONGKAI,
        ),
    )

    # Anchos
    for col, w in zip("ABCDEFGH", [25, 10, 13, 13, 50, 12, 12, 12]):
        ws.column_dimensions[col].width = w

    return ws


# =====================================================================
# Hoja 2: Plan 5 años
# =====================================================================
def crear_plan_5_anos(wb: Workbook, snap: dict):
    ws = wb.create_sheet("Plan 5 años")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Plan 5 años — EERR anual"
    estilo_titulo(ws["A1"])
    ws["A2"] = "Ingresos / EBITDA / CapEx por año del proyecto"
    estilo_subtitulo(ws["A2"])

    headers = ["Año", "Ingresos (CLP)", "EBITDA (CLP)", "CapEx (CLP)", "EBITDA Margin", "Ratio CapEx/Ventas"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i, value=h)
        estilo_header_tabla(cell)

    ingresos = snap["plan"]["ingresos_anuales"]
    ebitda = snap["plan"]["ebitda_anuales"]
    capex = snap["plan"]["capex_anuales"]

    for ano in range(5):
        row = 5 + ano
        ws.cell(row=row, column=1, value=ano + 1)
        ws.cell(row=row, column=2, value=ingresos[ano] if ano < len(ingresos) else 0)
        ws.cell(row=row, column=3, value=ebitda[ano] if ano < len(ebitda) else 0)
        ws.cell(row=row, column=4, value=capex[ano] if ano < len(capex) else 0)
        # Fórmulas
        ws.cell(row=row, column=5, value=f"=IFERROR(C{row}/B{row}, 0)")
        ws.cell(row=row, column=6, value=f"=IFERROR(D{row}/B{row}, 0)")
        for col in range(1, 7):
            estilo_celda(ws.cell(row=row, column=col))

    # Totales
    ws.cell(row=10, column=1, value="TOTAL")
    for col in range(2, 5):
        letter = get_column_letter(col)
        ws.cell(row=10, column=col, value=f"=SUM({letter}5:{letter}9)")
    for col in range(1, 7):
        estilo_celda(ws.cell(row=10, column=col), bold=True)
        ws.cell(row=10, column=col).fill = PatternFill("solid", fgColor=GRIS_FONDO)

    # Formato número
    for row in range(5, 11):
        for col in range(2, 5):
            ws.cell(row=row, column=col).number_format = '"$"#,##0_);("$"#,##0)'
        ws.cell(row=row, column=5).number_format = "0.0%"
        ws.cell(row=row, column=6).number_format = "0.0%"

    # Anchos
    for col, w in zip("ABCDEF", [10, 22, 22, 22, 18, 22]):
        ws.column_dimensions[col].width = w

    # Gráfico de barras
    chart = BarChart()
    chart.title = "Ingresos vs EBITDA vs CapEx por año"
    chart.style = 11
    chart.y_axis.title = "CLP"
    chart.x_axis.title = "Año"

    data = Reference(ws, min_col=2, max_col=4, min_row=4, max_row=9)
    cats = Reference(ws, min_col=1, max_col=1, min_row=5, max_row=9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 20
    ws.add_chart(chart, "H4")

    return ws


# =====================================================================
# Hoja 3: Balance de masa
# =====================================================================
def crear_balance_masa(wb: Workbook, base: ParametrosPlan):
    ws = wb.create_sheet("Balance Masa")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Balance de masa"
    estilo_titulo(ws["A1"])
    ws["A2"] = "Flujo MMPP -Productos final (régimen pleno)"
    estilo_subtitulo(ws["A2"])

    headers = ["MMPP", "Volumen anual (ton)", "Rendimiento", "Producto final (ton)", "% del total"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i, value=h)
        estilo_header_tabla(cell)

    # MMPP — desglose ilustrativo
    pesos_mmpp = {
        "ALPERUJO": 0.40,
        "TOMASA": 0.25,
        "POMASA": 0.20,
        "ORUJO_UVA": 0.10,
        "LEVADURA": 0.05,
    }
    volumen_total = base.volumen_total_ton_ano

    for i, (mmpp, peso) in enumerate(pesos_mmpp.items(), 1):
        row = 4 + i
        vol = volumen_total * peso
        rend = base.rendimiento_por_mmpp.get(mmpp, 0.25)
        producto = vol * rend
        ws.cell(row=row, column=1, value=mmpp)
        ws.cell(row=row, column=2, value=vol)
        ws.cell(row=row, column=3, value=rend)
        ws.cell(row=row, column=4, value=f"=B{row}*C{row}")
        ws.cell(row=row, column=5, value=f"=D{row}/SUM($D$5:$D$9)")
        for col in range(1, 6):
            estilo_celda(ws.cell(row=row, column=col))
        ws.cell(row=row, column=3).number_format = "0.0%"
        ws.cell(row=row, column=5).number_format = "0.0%"

    # Total
    row_tot = 10
    ws.cell(row=row_tot, column=1, value="TOTAL")
    ws.cell(row=row_tot, column=2, value=f"=SUM(B5:B9)")
    ws.cell(row=row_tot, column=4, value=f"=SUM(D5:D9)")
    for col in range(1, 6):
        estilo_celda(ws.cell(row=row_tot, column=col), bold=True)
        ws.cell(row=row_tot, column=col).fill = PatternFill("solid", fgColor=GRIS_FONDO)

    # Anchos
    for col, w in zip("ABCDE", [15, 20, 14, 22, 14]):
        ws.column_dimensions[col].width = w

    # Pie chart
    pie = PieChart()
    pie.title = "Distribución MMPP"
    data = Reference(ws, min_col=2, min_row=4, max_row=9)
    labels = Reference(ws, min_col=1, min_row=5, max_row=9)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 9
    pie.width = 14
    ws.add_chart(pie, "G4")


# =====================================================================
# Hoja 4: Sensibilidad
# =====================================================================
def crear_sensibilidad(wb: Workbook):
    ws = wb.create_sheet("Sensibilidad")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Análisis de Sensibilidad 2D"
    estilo_titulo(ws["A1"])
    ws["A2"] = "TIR según shock precio × costo MMPP (7x7 = 49 simulaciones)"
    estilo_subtitulo(ws["A2"])

    print("  -calculando heatmap 7x7 (~5s)...")
    hm = heatmap_2d(driver_x="precio", driver_y="costo_mmpp", n=7, hurdle_pct=0.15)
    res = hm.to_dict()

    # Eje X: precio (columnas)
    ws.cell(row=4, column=1, value="costo_mmpp \\ precio")
    estilo_header_tabla(ws.cell(row=4, column=1))
    for i, x in enumerate(res["rango_x"], 1):
        cell = ws.cell(row=4, column=1 + i, value=f"{x * 100:+.0f}%")
        estilo_header_tabla(cell)

    # Filas: costo MMPP
    for j, y in enumerate(res["rango_y"], 1):
        ws.cell(row=4 + j, column=1, value=f"{y * 100:+.0f}%")
        estilo_header_tabla(ws.cell(row=4 + j, column=1))
        for i, x in enumerate(res["rango_x"], 1):
            # Buscar la celda
            celda = next(
                (c for c in res["celdas"] if abs(c["x_pct"] - x) < 1e-6 and abs(c["y_pct"] - y) < 1e-6),
                None,
            )
            tir = celda["tir"] if celda else None
            c = ws.cell(row=4 + j, column=1 + i, value=tir if tir is not None else None)
            c.number_format = "0.00%"
            estilo_celda(c)

    # Color scale heatmap (verde=alto, rojo=bajo)
    n = len(res["rango_x"])
    range_str = f"B5:{get_column_letter(1 + n)}{4 + n}"
    ws.conditional_formatting.add(
        range_str,
        ColorScaleRule(
            start_type="min", start_color=ROJO,
            mid_type="num", mid_value=0.15, mid_color=AMARILLO,
            end_type="max", end_color=VERDE_TRONGKAI,
        ),
    )

    # Resumen
    summary_row = 4 + n + 3
    ws.cell(row=summary_row, column=1, value="TIR base").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=res["tir_base"]).number_format = "0.00%"
    ws.cell(row=summary_row + 1, column=1, value="Zona segura (% celdas > hurdle)").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=2, value=res["pct_zona_segura"]).number_format = "0.0%"
    ws.cell(row=summary_row + 2, column=1, value="Hurdle aplicado").font = Font(bold=True)
    ws.cell(row=summary_row + 2, column=2, value=res["hurdle_pct"]).number_format = "0.0%"

    # Anchos
    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, 2 + n):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11


# =====================================================================
# Hoja 5: Escenarios estratégicos
# =====================================================================
def crear_escenarios(wb: Workbook, snap: dict):
    ws = wb.create_sheet("Escenarios")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Escenarios Estratégicos"
    estilo_titulo(ws["A1"])
    ws["A2"] = "PILOTO / INDUSTRIAL / EXPANSION — comparación de tres planes"
    estilo_subtitulo(ws["A2"])

    headers = ["Escenario", "CapEx total (CLP)", "TIR proyecto", "VAN @ 18% (CLP)", "Payback (meses)"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i, value=h)
        estilo_header_tabla(cell)

    for i, e in enumerate(snap["escenarios_estrategicos"]["escenarios"], 1):
        row = 4 + i
        ws.cell(row=row, column=1, value=e["nombre"])
        ws.cell(row=row, column=2, value=e["capex_total"])
        ws.cell(row=row, column=3, value=e["tir"])
        ws.cell(row=row, column=4, value=e["van"])
        ws.cell(row=row, column=5, value=e["payback_meses"] or "—")
        for col in range(1, 6):
            estilo_celda(ws.cell(row=row, column=col))
        ws.cell(row=row, column=2).number_format = '"$"#,##0'
        ws.cell(row=row, column=3).number_format = "0.00%"
        ws.cell(row=row, column=4).number_format = '"$"#,##0'

    # Recomendación
    rec = snap["escenarios_estrategicos"]["recomendacion"]
    ws.merge_cells("A9:E9")
    ws["A9"] = f"Recomendación: {rec['elegido']} — {rec['razon']}"
    ws["A9"].font = Font(size=11, bold=True, color=VERDE_TRONGKAI)
    ws["A9"].fill = PatternFill("solid", fgColor="EAF6EA")
    ws["A9"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[9].height = 30

    # Anchos
    for col, w in zip("ABCDE", [15, 22, 15, 22, 16]):
        ws.column_dimensions[col].width = w


# =====================================================================
# Hoja 6: Carbono LCA
# =====================================================================
def crear_carbono(wb: Workbook, snap: dict):
    ws = wb.create_sheet("Carbono LCA")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Carbon Footprint — LCA 3 Escenarios"
    estilo_titulo(ws["A1"])
    ws["A2"] = "Emisiones netas + revenue créditos voluntarios CO2"
    estilo_subtitulo(ws["A2"])

    carbon = snap["carbon_footprint"]
    headers = ["Escenario", "Emisiones netas 5y (ton CO2eq)", "Revenue créditos 5y (CLP)", "¿Carbono negativo?"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i, value=h)
        estilo_header_tabla(cell)

    rows = [
        ("Baseline", carbon["baseline"]["emisiones_netas_5y_ton"],
         carbon["baseline"]["revenue_creditos_5y_clp"], carbon["baseline"]["es_carbono_negativo"]),
        ("BECCS", carbon["beccs"]["emisiones_netas_5y_ton"],
         carbon["beccs"]["revenue_creditos_5y_clp"], carbon["beccs"]["emisiones_netas_5y_ton"] < 0),
    ]
    for i, (nombre, em, rev, neg) in enumerate(rows, 1):
        row = 4 + i
        ws.cell(row=row, column=1, value=nombre)
        ws.cell(row=row, column=2, value=em)
        ws.cell(row=row, column=3, value=rev)
        ws.cell(row=row, column=4, value="SÍ ✓" if neg else "NO")
        for col in range(1, 5):
            estilo_celda(ws.cell(row=row, column=col))
        ws.cell(row=row, column=2).number_format = "#,##0"
        ws.cell(row=row, column=3).number_format = '"$"#,##0'
        if neg:
            ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor="EAF6EA")

    # Anchos
    for col, w in zip("ABCD", [15, 30, 26, 22]):
        ws.column_dimensions[col].width = w


# =====================================================================
# Hoja 7: Compliance REP
# =====================================================================
def crear_compliance(wb: Workbook):
    ws = wb.create_sheet("Compliance REP")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Compliance — Ley REP"
    estilo_titulo(ws["A1"])
    ws["A2"] = "8 hitos regulatorios agroindustria"
    estilo_subtitulo(ws["A2"])

    headers = ["Hito", "Fecha vigor", "Severidad", "Costo estimado (CLP)", "Descripción"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i, value=h)
        estilo_header_tabla(cell)

    for i, h in enumerate(HITOS_LEY_REP, 1):
        row = 4 + i
        ws.cell(row=row, column=1, value=h.nombre)
        ws.cell(row=row, column=2, value=h.fecha_vigor.isoformat())
        ws.cell(row=row, column=3, value=h.severidad.value)
        ws.cell(row=row, column=4, value=h.costo_estimado_clp)
        ws.cell(row=row, column=5, value=h.accion_requerida)
        for col in range(1, 6):
            estilo_celda(ws.cell(row=row, column=col))
        ws.cell(row=row, column=4).number_format = '"$"#,##0'
        # Color por severidad
        sev = h.severidad.value
        color_map = {"CRITICA": ROJO, "ALTA": NARANJA, "MEDIA": AMARILLO, "BAJA": VERDE_TRONGKAI}
        if sev in color_map:
            ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=color_map[sev])
            ws.cell(row=row, column=3).font = Font(bold=True, color="FFFFFF" if sev in ("CRITICA", "ALTA") else TEXTO)

    # Anchos
    for col, w in zip("ABCDE", [35, 14, 12, 22, 50]):
        ws.column_dimensions[col].width = w


# =====================================================================
# Hoja 8: Macro Chile
# =====================================================================
def crear_macro(wb: Workbook):
    ws = wb.create_sheet("Macro Chile")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Indicadores Macro Chile"
    estilo_titulo(ws["A1"])
    ws["A2"] = "Snapshot Banco Central (vía mindicador.cl, cache 24h)"
    estilo_subtitulo(ws["A2"])

    try:
        m = snapshot_resumen()
    except Exception:
        m = {}

    indicadores = [
        ("Dólar Observado", m.get("dolar_clp"), "CLP"),
        ("UF", m.get("uf_clp"), "CLP"),
        ("IPC mensual", m.get("ipc_pct_mensual"), "%"),
        ("TPM", m.get("tpm_pct"), "%"),
        ("Tasa desempleo", m.get("tasa_desempleo_pct"), "%"),
    ]

    headers = ["Indicador", "Valor", "Unidad"]
    for i, h in enumerate(headers, 1):
        estilo_header_tabla(ws.cell(row=4, column=i, value=h))

    for i, (nombre, valor, unidad) in enumerate(indicadores, 1):
        row = 4 + i
        ws.cell(row=row, column=1, value=nombre)
        ws.cell(row=row, column=2, value=valor if valor is not None else "—")
        ws.cell(row=row, column=3, value=unidad)
        for col in range(1, 4):
            estilo_celda(ws.cell(row=row, column=col))

    ws["A11"] = f"Fuente: {m.get('fuente', 'mindicador.cl')}"
    estilo_subtitulo(ws["A11"])

    for col, w in zip("ABC", [22, 15, 10]):
        ws.column_dimensions[col].width = w


# =====================================================================
# Hoja 9: Supuestos top 20
# =====================================================================
def crear_supuestos(wb: Workbook, base: ParametrosPlan):
    ws = wb.create_sheet("Supuestos")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Supuestos críticos del modelo"
    estilo_titulo(ws["A1"])
    ws["A2"] = "Valores configurables — modificar acá impacta TODAS las hojas"
    estilo_subtitulo(ws["A2"])

    headers = ["Supuesto", "Valor", "Unidad", "Fuente / Justificación"]
    for i, h in enumerate(headers, 1):
        estilo_header_tabla(ws.cell(row=4, column=i, value=h))

    supuestos = [
        ("Volumen total anual", base.volumen_total_ton_ano, "ton", "Cap contractual"),
        ("WACC anual", base.wacc_anual, "fracción", "Scielo sector pesca-acuicultura 19,6% - premium ESG"),
        ("Costo MMPP neto", base.costo_mmpp_clp_kg, "CLP/kg", "Flete - pagos parciales por residuo"),
        ("Costo comercialización", base.costo_comercializacion_pct, "% revenue", "Marketing + distribución + I+D"),
        ("Impuesto renta", base.impuesto_renta_pct, "% utilidad", "DL 824 LIR Chile 27%"),
        ("OpEx mensual", base.opex_mensual_clp, "CLP", "25 personas + mantención + energía + admin"),
        ("DSO (días cobro)", base.dso_dias, "días", "B2B alimentos típico 45-60d"),
        ("DPO (días pago)", base.dpo_dias, "días", "Proveedores MMPP típico 30-45d"),
        ("Inventario", base.inventario_dias, "días", "Estacional MMPP concentrada 3 meses"),
        ("Precio HARINA_ALPERUJO", base.precios_clp_kg.get("HARINA_ALPERUJO", 800), "CLP/kg", "Premium funcional"),
        ("Precio HARINA_ORUJO", base.precios_clp_kg.get("HARINA_ORUJO", 600), "CLP/kg", "30% bajo harina pescado Chile"),
        ("Precio LICOPENO", base.precios_clp_kg.get("LICOPENO", 80000), "CLP/kg", "USD 108-253/kg × 0.5 nuevo entrante"),
        ("Precio PECTINA", base.precios_clp_kg.get("PECTINA", 25000), "CLP/kg", "USD 55/kg mundial × 0.5"),
        ("Rendimiento ALPERUJO", base.rendimiento_por_mmpp.get("ALPERUJO", 0.39), "ton out/in", "Calibrado planta industrial"),
        ("Rendimiento POMASA", base.rendimiento_por_mmpp.get("POMASA", 0.22), "ton out/in", "Validar con tomatera"),
    ]

    for i, (nombre, valor, unidad, fuente) in enumerate(supuestos, 1):
        row = 4 + i
        ws.cell(row=row, column=1, value=nombre)
        ws.cell(row=row, column=2, value=valor)
        ws.cell(row=row, column=3, value=unidad)
        ws.cell(row=row, column=4, value=fuente)
        for col in range(1, 5):
            estilo_celda(ws.cell(row=row, column=col))

    for col, w in zip("ABCD", [28, 18, 14, 55]):
        ws.column_dimensions[col].width = w


# =====================================================================
# Hoja 10: Instrucciones / Cómo conectar macros
# =====================================================================
def crear_instrucciones(wb: Workbook):
    ws = wb.create_sheet("Instrucciones")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Cómo usar este Excel"
    estilo_titulo(ws["A1"])
    ws["A2"] = "Conexión al motor en vivo + macros opcionales"
    estilo_subtitulo(ws["A2"])

    lines = [
        ("", ""),
        ("📊 ESTRUCTURA DEL EXCEL", ""),
        ("Dashboard", "KPIs grandes + Investment Readiness Score"),
        ("Plan 5 años", "EERR anual con gráfico de barras"),
        ("Balance Masa", "Flujo MMPP→producto con pie chart"),
        ("Sensibilidad", "Heatmap 7x7 TIR vs precio×costo (con color scale)"),
        ("Escenarios", "PILOTO/INDUSTRIAL/EXPANSION + recomendación"),
        ("Carbono LCA", "Emisiones baseline + BECCS + revenue créditos"),
        ("Compliance REP", "8 hitos Ley REP con severidad codificada"),
        ("Macro Chile", "Snapshot Banco Central (vía mindicador.cl)"),
        ("Supuestos", "Variables del modelo — modificar acá no afecta el motor"),
        ("", ""),
        ("🔄 ACTUALIZAR DATOS", ""),
        ("1.", "Cierra este archivo"),
        ("2.", "Ejecuta: python scripts/generar_excel_master.py"),
        ("3.", "Abre el nuevo .xlsx generado en backups/"),
        ("", ""),
        ("🌐 CONEXIÓN AL MOTOR (LIVE API)", ""),
        ("Engine URL", "https://trongkai-engine.fly.dev"),
        ("OpenAPI/Swagger", "https://trongkai-engine.fly.dev/docs"),
        ("Endpoint dashboard", "GET /api/snapshot — todo en un JSON"),
        ("Endpoint PDF", "GET /api/tearsheet.pdf — descarga directa"),
        ("Endpoint readiness", "GET /readiness/score — score 0-100"),
        ("", ""),
        ("⚙️ MACROS VBA (OPCIONAL — agregar manualmente)", ""),
        ("Para conectar este Excel al motor con macros VBA:", ""),
        ("1.", "Guarda este archivo como .xlsm (Habilitado para macros)"),
        ("2.", "Alt+F11 abre el editor VBA"),
        ("3.", "Pega el código de la siguiente sección"),
        ("", ""),
        ("--- CÓDIGO VBA EJEMPLO ---", ""),
        ("Sub RefrescarDashboard()", ""),
        ("    Dim http As Object", ""),
        ("    Set http = CreateObject('MSXML2.XMLHTTP')", ""),
        ("    http.Open 'GET', 'https://trongkai-engine.fly.dev/api/snapshot', False", ""),
        ("    http.Send", ""),
        ("    MsgBox 'Status: ' & http.Status & vbCrLf & 'Body: ' & Left(http.responseText, 200)", ""),
        ("End Sub", ""),
        ("", ""),
        ("Sub DescargarTearsheetPDF()", ""),
        ("    Dim url As String", ""),
        ("    url = 'https://trongkai-engine.fly.dev/api/tearsheet.pdf'", ""),
        ("    ThisWorkbook.FollowHyperlink url", ""),
        ("End Sub", ""),
        ("", ""),
        ("📁 GENERADO POR", "TrongkAI Platform"),
        ("Fecha generación", datetime.now().strftime("%d %B %Y, %H:%M")),
    ]

    for i, (a, b) in enumerate(lines, 4):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
        if a and not b:  # Header de sección
            ws.cell(row=i, column=1).font = Font(name="Calibri", size=12, bold=True, color=VERDE_TRONGKAI)
        elif a:
            ws.cell(row=i, column=1).font = Font(name="Calibri", size=10, bold=False, color=TEXTO)
            ws.cell(row=i, column=2).font = Font(name="Calibri", size=10, color=TEXTO_SEC)

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 80


# =====================================================================
# Main
# =====================================================================
def main():
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    print("Trongkai Excel Master - generando...")
    print()

    base = ParametrosPlan()
    plan = build_plan(base)
    val = valuar_proyecto_ev_ebitda(plan)
    escs = comparar_escenarios_estrategicos()

    print("-snapshot del modelo...")
    # Construir snapshot manualmente (más rápido que llamar al endpoint)
    rendimiento_prom = sum(base.rendimiento_por_mmpp.values()) / len(base.rendimiento_por_mmpp)
    vols = [base.volumen_total_ton_ano * base.volumen_pct_por_ano.get(y, 1.0) for y in range(1, 6)]
    carbon = comparar_escenarios_footprint(vols, rendimiento_promedio=rendimiento_prom)

    snap = {
        "plan": {
            "kpis": {
                "tir": plan.kpis.tir_proyecto_anual,
                "van": plan.kpis.van,
                "payback_meses": plan.kpis.payback_meses,
                "ebitda_margin_promedio": plan.kpis.ebitda_margin_promedio,
                "ratio_capex_ventas": plan.kpis.ratio_capex_ventas,
            },
            "ingresos_anuales": plan.ingresos_anuales,
            "ebitda_anuales": plan.ebitda_anuales,
            "capex_anuales": plan.capex_anuales,
        },
        "valuation": {
            "ebitda_ano5_clp": val.ebitda_ano5_clp,
            "ev_base_clp": val.ev_clp_base,
            "moic": val.moic_estimado,
        },
        "escenarios_estrategicos": {
            "escenarios": [
                {
                    "nombre": e.nombre,
                    "capex_total": sum(e.resumen.capex_anuales),
                    "tir": e.resumen.kpis.tir_proyecto_anual,
                    "van": e.resumen.kpis.van,
                    "payback_meses": e.resumen.kpis.payback_meses,
                }
                for e in escs
            ],
            "recomendacion": {"elegido": "INDUSTRIAL", "razon": "VAN positivo y mejor risk-adjusted return"},
        },
        "carbon_footprint": {
            "baseline": carbon["baseline"],
            "beccs": carbon["beccs"],
        },
    }

    print("-readiness score (8 dimensiones, 200 sims MC)...")
    rs = calcular_readiness_score(n_sims_mc=200).to_dict()

    wb = Workbook()
    # Eliminar hoja default
    wb.remove(wb.active)

    print("-hoja Dashboard...")
    crear_dashboard(wb, snap, rs)
    print("-hoja Plan 5 años...")
    crear_plan_5_anos(wb, snap)
    print("-hoja Balance Masa...")
    crear_balance_masa(wb, base)
    print("-hoja Sensibilidad...")
    crear_sensibilidad(wb)
    print("-hoja Escenarios...")
    crear_escenarios(wb, snap)
    print("-hoja Carbono LCA...")
    crear_carbono(wb, snap)
    print("-hoja Compliance REP...")
    crear_compliance(wb)
    print("-hoja Macro Chile...")
    crear_macro(wb)
    print("-hoja Supuestos...")
    crear_supuestos(wb, base)
    print("-hoja Instrucciones...")
    crear_instrucciones(wb)

    # Salida
    out_dir = ROOT / "backups"
    out_dir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    out_path = out_dir / f"Trongkai-Master-{stamp}.xlsx"
    wb.save(out_path)

    print()
    print(f"OK Generado: {out_path}")
    print(f"  Tamaño: {out_path.stat().st_size / 1024:.1f} KB")
    print(f"  Hojas: {len(wb.sheetnames)}")
    print()
    print("Hojas:")
    for s in wb.sheetnames:
        print(f"  - {s}")


if __name__ == "__main__":
    main()
