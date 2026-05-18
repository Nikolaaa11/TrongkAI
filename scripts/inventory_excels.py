"""Inventario de los 3 Excels del cliente Trongkai.

Volca contenido a docs/INVENTARIO-EXCELS.md y produce un dump de cada hoja
para análisis posterior. Lee siguiendo las reglas del skill xlsx
(openpyxl, lectura de valores y fórmulas).
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

CTX = Path("C:/Users/nicol/OneDrive/Documentos/0.1.1 TrongkAI/trongkai-platform/contexto")
OUT = Path("C:/Users/nicol/OneDrive/Documentos/0.1.1 TrongkAI/trongkai-platform/docs")

FILES = [
    "Info_Plan_5_anos_Estructura_A.xlsx",
    "Cuadro_PPTO_Variables_PD_Plan_5_Anos_A.xlsx",
    "Tareas_Plan_5_anos.xlsx",
]


def inventory_sheet(ws) -> dict:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    non_empty = 0
    pd_count = 0
    ok_count = 0
    formula_count = 0
    sample = []
    for row in ws.iter_rows(min_row=1, max_row=min(max_row, 200), values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            non_empty += 1
            if isinstance(v, str):
                up = v.strip().upper()
                if up == "PD":
                    pd_count += 1
                elif up == "OK":
                    ok_count += 1
                if v.startswith("="):
                    formula_count += 1
            if len(sample) < 15 and v not in (None, ""):
                sample.append({"cell": cell.coordinate, "value": str(v)[:120]})
    return {
        "dims": f"{max_row}x{max_col}",
        "non_empty": non_empty,
        "PD": pd_count,
        "OK": ok_count,
        "formula_strings": formula_count,
        "sample": sample,
    }


def dump_sheet(ws, target: Path) -> None:
    """Dump full sheet content as tab-separated text for human review."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        cleaned = ["" if v is None else str(v).replace("\t", " ").replace("\n", " | ") for v in row]
        if any(c != "" for c in cleaned):
            rows.append("\t".join(cleaned))
    target.write_text("\n".join(rows), encoding="utf-8")


def main() -> None:
    report = {}
    for fname in FILES:
        path = CTX / fname
        print(f"\n=== {fname} ===", file=sys.stderr)
        wb = openpyxl.load_workbook(path, data_only=False)
        wb_values = openpyxl.load_workbook(path, data_only=True)
        sheets = []
        for sname in wb.sheetnames:
            ws = wb[sname]
            ws_val = wb_values[sname]
            info = inventory_sheet(ws)
            info["values_dims"] = f"{ws_val.max_row}x{ws_val.max_column}"
            sheets.append({"name": sname, **info})
            safe = sname.replace("/", "_").replace("\\", "_")
            dump_dir = OUT / "dumps" / fname.replace(".xlsx", "")
            dump_dir.mkdir(parents=True, exist_ok=True)
            dump_sheet(ws_val, dump_dir / f"{safe}.tsv")
        report[fname] = {"sheets": sheets}

    (OUT / "_inventory.json").write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print("\nWrote", OUT / "_inventory.json")


if __name__ == "__main__":
    main()
